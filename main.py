import asyncio
import json
from dataclasses import dataclass
from datetime import datetime
from decimal import Decimal
from typing import List
import time

import tomli
from curl_cffi.requests import AsyncSession
from loguru import logger
from openpyxl.styles import Side, Border, Font, Alignment, PatternFill
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class NetworkBalance:
    network: str
    balance: Decimal
    token: str = "ETH"

    def __str__(self) -> str:
        return f"{self.network}: {self.balance:.6f} {self.token}"


class WalletBalances:
    def __init__(self, address: str):
        self.address = address
        self.balances: List[NetworkBalance] = []

    def add_balance(self, network: str, balance: Decimal, token: str = "ETH"):
        self.balances.append(NetworkBalance(network, balance, token))


class BalanceReport:
    def __init__(self):
        self.wallets: List[WalletBalances] = []

    def add_wallet(self, wallet: WalletBalances):
        self.wallets.append(wallet)

# Глобальный семафор для ограничения количества одновременных запросов
REQUEST_SEMAPHORE = asyncio.Semaphore(3)  # Максимум 3 одновременных запроса

# Словарь для хранения времени последнего запроса к каждой сети
last_request_time = {}
# Минимальный интервал между запросами к одной сети (в секундах)
MIN_REQUEST_INTERVAL = 3.0

# Словарь для отслеживания текущего индекса RPC для каждой сети
current_rpc_index = {}

def load_networks_config():
    """Load network RPC configurations and thresholds from config.toml file."""
    try:
        with open("config.toml", "rb") as f:
            config = tomli.load(f)
            return {
                "networks": {network: rpcs["rpcs"] for network, rpcs in config["networks"].items()},
                "thresholds": config["thresholds"]
            }
    except Exception as e:
        logger.error(f"Error loading config.toml: {e}")
        raise

# Загрузка конфигурации сетей и пороговых значений из TOML файла
config = load_networks_config()
networks = config["networks"]
thresholds = config["thresholds"]

headers = {
    'Content-Type': 'application/json',
    'Accept': 'application/json'
}

def get_next_rpc(network: str) -> str:
    """Получает следующий доступный RPC для сети."""
    if network not in current_rpc_index:
        current_rpc_index[network] = 0
    
    rpc_list = networks[network]
    current_index = current_rpc_index[network]
    
    # Переключаемся на следующий RPC
    next_index = (current_index + 1) % len(rpc_list)
    current_rpc_index[network] = next_index
    
    return rpc_list[current_index]

async def wait_for_rate_limit(network: str):
    """Ждем необходимое время перед следующим запросом к сети."""
    current_time = time.time()
    if network in last_request_time:
        elapsed = current_time - last_request_time[network]
        if elapsed < MIN_REQUEST_INTERVAL:
            await asyncio.sleep(MIN_REQUEST_INTERVAL - elapsed)
    last_request_time[network] = time.time()


async def get_balance_for_network(session: AsyncSession, address: str, network: str) -> Decimal:
    """Get balance for a specific network using curl_cffi."""
    retries = len(networks[network])  # Количество попыток равно количеству доступных RPC
    last_error = None
    rpc_url = networks[network][0]

    for _ in range(retries):
        try:
            payload = {
                "jsonrpc": "2.0",
                "method": "eth_getBalance",
                "params": [address, "latest"],
                "id": 1
            }

            async with REQUEST_SEMAPHORE:
                await wait_for_rate_limit(network)
                
                response = await session.post(
                    rpc_url,
                    data=json.dumps(payload),
                    headers=headers,
                    timeout=10
                )

                if response.status_code == 429 or "rate limit" in response.text.lower():
                    # Если достигнут лимит запросов, переключаемся на следующий RPC
                    rpc_url = get_next_rpc(network)
                    logger.warning(f"Rate limit reached for {network}, switching to next RPC: {rpc_url}")
                    continue

                if response.status_code == 200:
                    result = response.json()
                    if "error" in result:
                        if "rate" in str(result["error"]).lower():
                            # Если в ответе есть ошибка о превышении лимита
                            rpc_url = get_next_rpc(network)
                            logger.warning(f"Rate limit error for {network}, switching to next RPC: {rpc_url}")
                            continue
                        raise Exception(f"RPC Error: {result['error']}")
                    
                    balance = int(result["result"], 16)
                    return Decimal(balance) / Decimal(10**18)
                
                response.raise_for_status()

        except Exception as e:
            last_error = e
            logger.error(f"Error getting balance for {network} ({rpc_url}): {str(e)}")
            # Переключаемся на следующий RPC при любой ошибке
            rpc_url = get_next_rpc(network)
            logger.info(f"Switching to next RPC for {network}: {rpc_url}")
            continue

    # Если все RPC исчерпаны, возвращаем 0 и логируем ошибку
    logger.error(f"All RPCs failed for {network}. Last error: {last_error}")
    return Decimal(0)


async def process_wallet_chunk(wallets: List[str], session: AsyncSession) -> List[WalletBalances]:
    """Process a chunk of wallets."""
    wallet_balances_list = []

    for wallet in wallets:
        try:
            # address = Web3.to_checksum_address(wallet.strip())
            address = wallet.strip()
            wallet_balances = WalletBalances(address)

            # Создаем список корутин для всех сетей
            balance_tasks = [
                get_balance_for_network(session, address, network_name)
                for network_name in networks.keys()
            ]

            # Запускаем все запросы параллельно
            balances = await asyncio.gather(*balance_tasks)

            # Добавляем результаты
            for (network_name, _), balance in zip(networks.items(), balances):
                coin_symbol = 'BNB' if network_name == 'BSC' else ('POL' if network_name == 'Polygon' else 'ETH')

                if balance > 0:
                    logger.info(f"Found balance in {network_name}: {address} = {balance} {coin_symbol}")
                wallet_balances.add_balance(network_name, balance, token=coin_symbol)

            wallet_balances_list.append(wallet_balances)

        except Exception as e:
            logger.error(f"Error processing wallet {wallet}: {e}")
            wallet_balances = WalletBalances(wallet)
            wallet_balances.add_balance("ERROR", Decimal(0), f"Error: {str(e)}")
            wallet_balances_list.append(wallet_balances)

    return wallet_balances_list


async def evm_checker(wallets: list[str]) -> BalanceReport:
    """Check EVM balances for multiple wallets in parallel."""
    report = BalanceReport()

    # Разбиваем список кошельков на 5 частей
    chunk_size = (len(wallets) + 4) // 5  # округление вверх
    wallet_chunks = [wallets[i:i + chunk_size] for i in range(0, len(wallets), chunk_size)]
    logger.info(f"Processing {len(wallets)} wallets in {len(wallet_chunks)} chunks")

    async with AsyncSession() as session:
        # Создаем задачи для каждого чанка
        tasks = [process_wallet_chunk(chunk, session) for chunk in wallet_chunks]

        # Запускаем все чанки параллельно
        chunks_results = await asyncio.gather(*tasks)

        # Собираем результаты
        total_processed = 0
        for chunk_result in chunks_results:
            for wallet_balances in chunk_result:
                report.add_wallet(wallet_balances)
                total_processed += 1

        logger.info(f"Successfully processed {total_processed} wallets")

    return report


def write_results(report: BalanceReport):
    wb = Workbook()
    sheet: Worksheet = wb.active

    now = datetime.now().replace(microsecond=0)
    filename = f".\\EVM_report_{now.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

    sheet.name = now
    for column in sheet.iter_cols(min_col=1, max_col=1):
        sheet.column_dimensions[column[0].column_letter].width = 45
    
    # Установка фиксированной ширины для всех столбцов
    for column in sheet.iter_cols(min_col=2, max_col=11):
        sheet.column_dimensions[column[0].column_letter].width = 35  # Увеличиваем ширину для больших чисел

    # Стиль заголовков
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill("solid", fgColor="FFFF00")  # Желтый фон fgColor="FFFF00"

    # Создаем стиль для границ (все стороны)
    thin = Side(border_style="thin", color="000000")  # Тонкая черная линия
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Стиль для числовых значений
    number_alignment = Alignment(horizontal="right", vertical="center")
    number_format = "0." + "0" * 18  # Формат с 18 знаками после запятой

    # Стиль для подсветки значений выше порога
    green_fill = PatternFill("solid", fgColor="90EE90")  # Светло-зеленый цвет

    headers_xlsx = ['Address', 'Ethereum', 'Sepolia', 'Linea', 'Polygon',
               'Base', 'Blast', 'Optimism', 'Arbitrum', 'Zksync', 'BSC']

    # Записываем заголовки
    for col_num, header in enumerate(headers_xlsx, start=1):
        cell = sheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = header_alignment
        cell.fill = header_fill
        cell.border = border

    for wallet in report.wallets:
        ethereum_balance = sepolia_balance = linea_balance = polygon_balance = base_balance = blast_balance\
            = optimism_balance = arbitrum_balance = zksync_balance = bsc_balance = 'None'

        for balance in wallet.balances:
            if balance.network == 'Ethereum':
                ethereum_balance = balance.balance
            elif balance.network == 'Sepolia':
                sepolia_balance = balance.balance
            elif balance.network == 'Linea':
                linea_balance = balance.balance
            elif balance.network == 'Polygon':
                polygon_balance = balance.balance
            elif balance.network == 'Base':
                base_balance = balance.balance
            elif balance.network == 'Blast':
                blast_balance = balance.balance
            elif balance.network == 'Optimism':
                optimism_balance = balance.balance
            elif balance.network == 'Arbitrum':
                arbitrum_balance = balance.balance
            elif balance.network == 'Zksync':
                zksync_balance = balance.balance
            elif balance.network == 'BSC':
                bsc_balance = balance.balance

        # Определяем номер строки для записи (следующая свободная строка)
        row_num = sheet.max_row + 1

        # Записываем адрес
        address_cell = sheet.cell(row=row_num, column=1, value=wallet.address)
        address_cell.border = border
        address_cell.alignment = Alignment(horizontal="left", vertical="center")

        # Функция для установки стиля ячейки с балансом
        def set_balance_cell_style(cell, value, network):
            cell.border = border
            if value != 'None':
                cell.number_format = number_format
                cell.alignment = number_alignment
                cell.value = float(value)  # Преобразуем Decimal в float для Excel
                
                # Проверяем, превышает ли значение порог и подсвечиваем ячейку
                if network.lower() in thresholds and float(value) >= thresholds[network.lower()]:
                    cell.fill = green_fill
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.value = value

        # Записываем балансы с форматированием
        balances_with_networks = list(zip(
            [ethereum_balance, sepolia_balance, linea_balance, polygon_balance, 
             base_balance, blast_balance, optimism_balance, arbitrum_balance, 
             zksync_balance, bsc_balance],
            ['ethereum', 'sepolia', 'linea', 'polygon', 'base', 'blast', 
             'optimism', 'arbitrum', 'zksync', 'bsc']
        ))
        
        for col_num, (balance, network) in enumerate(balances_with_networks, start=2):
            cell = sheet.cell(row=row_num, column=col_num)
            set_balance_cell_style(cell, balance, network)

    wb.save(filename)
    wb.close()
    logger.info(f"Results saved to {filename}")


async def main():
    with open('wallets.txt') as f:
        wallets = f.readlines()
        wallets = [wallet.strip('\n') for wallet in wallets]

    report = await evm_checker(wallets)
    write_results(report)


if __name__ == '__main__':
    asyncio.run(main())
